import openpyxl as xl
import os
import json

# Get file path
filepath = input("Please type in the document name (Press RETURN to default to 'TT Brothers Information')") or "TT Brothers Information"
path = os.getcwd()
filepath = path + "/" + filepath + ".xlsx"

# Load source Excel sheet
wb1 = xl.load_workbook(filepath)
ws1 = wb1.worksheets[0]

# Get total rows/cols
maxRow = ws1.max_row
maxCol = ws1.max_column

# Set up JSON dict
json_data = {"brothers": []}

# Rip data and add to JSON dict
#   Headers: ID, Name (First Last), Class, Active (Y or N), LinkedIn, Major, Cabby/Exec (Y or N), Profile URL
for r in range(2, maxRow + 1):
    brother = {}
    for c in range(1, maxCol + 1):
        if c == 1:
            brother["id"] = ws1.cell(row=r, column=c).value
        if c == 2:
            brother["name"] = ws1.cell(row=r, column=c).value
        if c == 3:
            brother["class"] = ws1.cell(row=r, column=c).value
        if c == 4:
            brother["active_status"] = ws1.cell(row=r, column=c).value
        if c == 5:
            brother["linkedin_url"] = ws1.cell(row=r, column=c).value
        if c == 6:
            brother["major"] = ws1.cell(row=r, column=c).value
        if c == 7:
            brother["cabby_exec_status"] = ws1.cell(row=r, column=c).value
        if c == 8:
            brother["profile_url"] = ws1.cell(row=r, column=c).value
        if c > 8:
            raise Exception("ERROR: Column is " + str(c) + ", Cell Value is " + str(ws1.cell(row=r, column=c).value))
    json_data["brothers"].append(brother)

# Write JSON file
with open("brother_info.json", "w") as outfile:
    json.dump(json_data, outfile, indent=4)
